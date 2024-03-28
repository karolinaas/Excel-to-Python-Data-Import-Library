import numpy as np
from openpyxl import load_workbook

def import_data_column(filename: str, sheetname: str, row_pos: int, col_pos: int , number_of_rows: int):
    """ Return imported data
    Import a specified number of rows from a single column from an Excel spreadsheet and return their contents as an Numpy array.

    Parameters
    ------------
        filename: str
            Filename, including .xlsx extension. Absolute path can also be used.
        sheetname: str
            Name of the sheet where data is located
        row_pos: int
            Number of the row where the first cell to be imported is located.
        col_pos: int
            Number of the column where the first cell to be imported is located.
        number_of_rows: int
            Number of rows (cells in a column) to be imported.
    Return
    -----------
        data : ndarray
            Numpy array of imported data. 
    """

    # Open a workbook (file) and the specified worksheet
    wb = load_workbook(filename=filename, data_only=True)
    ws = wb[sheetname]

    # Create and empty Numpy array for spreadsheet data to be loaded into
    data = np.empty(number_of_rows)

    # Iterate over the specified rows in the worksheet and load its values into Numpy array
    i = 0
    for row in ws.iter_rows(min_row=row_pos, max_row=row_pos+number_of_rows-1, min_col=col_pos, max_col=col_pos, values_only=True):
        for value in row:
            data[i] = value
            i += 1

    return data
